import jieba
import xlwt
import os
import re
import xlrd
import pandas as pd
from  pandas import DataFrame
from openpyxl import load_workbook

from xlutils.copy import copy

stopwords = {}.fromkeys(['l','t','co','  ',"i","my","me",'you','your','yours','thy','their','them','we','our','it','is','am','are','be','was','were','will',
"would",'what','would','why','when','where','there','only','up','and','from','that','with','this','can','have','has','as','of','in','at','all',
"just",'about','an','one','also','only','been','some','way','which','then','on','for','https','or','only','if','than','the','but','just','like',
'who','over','more','about','do','no','see','out','done','its','to','too','any','these','after','below','off','did','a','think','into','many','few',
'everyone','could','so','because','done','his','her','him','something','let',"let's",'yet','such','people','the','and','is', "'", 'is', ' '])


industry_words = {}.fromkeys(['Stocks', 'futures', 'commodities', 'blockchain', 'bitcoin', 'crypto', 'synthetic', 'DeFi', 'ETH', 'funds', 'TradFi'])
first_kind_words = {}.fromkeys(['Unbanked', 'Democratization of Finance', 'borderless', 'decentralization', 'liberal', 'financial equality'])
second_kind_words = {}.fromkeys(['Moon', 'APY', 'Farm','bull', 'profit', 'airdrop'])

stock={}
execl = xlwt.Workbook(encoding='utf-8',style_compression=0)
kol_kind={}
execlKOL  = copy(xlrd.open_workbook(os.getcwd()+'\关键词得到的KOL列表0.xlsx'))
Kolsheet=execlKOL.add_sheet('all',cell_overwrite_ok=True)
t=1
sheet_len=0
#输出地址
outputdir = os.getcwd()

#设定KOL_list长度
def set_sheet_len(len):
    global sheet_len
    sheet_len=len

#处理下KOLList 弃用勿删，后续可能需要启用
def handle_main_excel():
    sheets = execlKOL.sheet_names()
    # print sheets

    # 循环遍历所有sheet
    alldata = DataFrame()
    for i in range(len(sheets)):
        df = pd.read_excel(excel_name, sheet_name=i, index=False, encoding='utf8')
        alldata = alldata.append(df)
    writer = pd.ExcelWriter(os.getcwd()+'\关键词结果.xls',engin='openpyxl')
    book = load_workbook(writer.path)
    writer.book = book
    #利用dataframe.to_excel保存合并后的数据到新的sheet
    alldata.to_excel(excel_writer=writer,sheet_name="all")#生成新的sheet命名为ALLDATA
    writer.save()
    writer.close()
    print ('处理完成！')



#统计关键字出现的频率
#@post 推文正文（拼接过的）
#@username 该推文对应的博主
#统计关键字出现的频率
def seprate(post,username):
    # 精确模式：试图将句子最精确地切开，适合文本分析；
    stock={}
    seg_list = jieba.cut(post, cut_all=False)
    for se in seg_list:

        # 
        # 正则去除特殊符号，后续可添加表情符号
        if se  in stopwords or se in '~!@#$%^&*()_\-+=<>?:"|,.\/;\\[\]·~！@#￥%……&*（）——\-+=|《》？：“”【】、；‘’，。、' or se.strip()=='':
            continue
        if se in stock:
            stock[se]+=1
        else:
            stock[se]=1

    global t
    
    # 统计表用户名写入
    Kolsheet.write(t,0,username) 
    #
    exportgNToExcl(stock,username)
    
    t+=1

# 保存分析结果到Kolsheet
# 保存
def exportgNToExcl(scs,fileName):

        sheet = execl.add_sheet(fileName,cell_overwrite_ok=True)
        i=1
        sheet.write(0,0,'word') 
        sheet.write(0,1,'frequency') 
        induss=''
        ca1=''
        ca2=''

        for key in scs: 
            if key  in industry_words and scs[key]>0:
                induss=induss+(key+":"+str(scs[key]))
                # print induss
                # kol_kind['industry']=1
            if key  in first_kind_words and scs[key]>0:
                ca1+=(key+":"+str(scs[key]))
                Kolsheet.write(t,2,key) 
                Kolsheet.write(t,3,'1') 
                Kolsheet.write(t,4,scs[key]) 
                
            if key  in second_kind_words and scs[key]>0:
                ca2+=(key+":"+str(scs[key]))
                Kolsheet.write(t,2,key) 
                Kolsheet.write(t,3,'2') 
                Kolsheet.write(t,4,scs[key]) 
            
            Kolsheet.write(t,1,induss) 
            Kolsheet.write(t,5,ca1) 
            Kolsheet.write(t,6,ca2) 
            
            if scs[key]<0 :
                continue
            
            # if ctweet.get('favorite_count')<10:
            #     continue

            sheet.write(i,0,key) 
            sheet.write(i,1,scs[key]) 
            i+=1
            # if t>sheet_len-1:

        #保证一定存储 -1
        if t>sheet_len-1:
            print("开始保存文件")
            execlKOL.save(outputdir+'/'+'KOLend4'+'.xls')
            execl.save(outputdir+'/'+'每个KOL的post高频词4'+'.xls')

def exportENDkolExcl(execlKOL):
    # handle_main_excel()
    # Kolsheet=execlKOL.sheet_by_name('all')
    Kolsheet.write(0,0,'账号') 
    Kolsheet.write(0,1,'industry') 
    Kolsheet.write(0,2,'cause') 
    Kolsheet.write(0,3,'type') 
    Kolsheet.write(0,4,'cause_frequency') 
    Kolsheet.write(0,5,'cause_one') 
    Kolsheet.write(0,6,'cause_two') 
    execlKOL.save(outputdir+'/'+'KOLend'+'.xls')

# def exportgToExcl(scs,fileName):
    
#         sheet = execl.add_sheet(fileName,cell_overwrite_ok=True)
#         i=1
#         sheet.write(0,0,'word') 
#         sheet.write(0,1,'frequency') 

#         for key in scs: 
            
#             if scs[key]<0 :
#                 continue
            
#             # if ctweet.get('favorite_count')<10:
#             #     continue

#             sheet.write(i,0,key) 
#             sheet.write(i,1,scs[key]) 
#             i+=1
#         execl.save(outputdir+'/'+'post高频词'+'.xls')

if __name__ == '__main__':
    seprate('Press any key to continue ')
